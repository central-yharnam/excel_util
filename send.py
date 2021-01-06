from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showerror
from functools import partial
import openpyxl
import os

# A utility with a GUI that transfers designated cells from one excel sheet to another.
# There are many excel sheets where the values themselves cannot be copy-pasted because they have formulas.
# Since they cannot be copypasted they must be written in manually, creating a lot of room for error.

class interface:
	

	def __init__(self):
		self.d_text_inp = None
		self.r_text_inp = None

		self.donor_range = None
		self.recipient_range = None

		self.donor_workbook = None
		self.donor_sheet = None

		self.recipient_workbook = None
		self.recipient_sheet = None
		
		self.output_fname = None

		self.interf = Tk()
		self.interf.grid_rowconfigure(0, weight=1)
		self.interf.grid_columnconfigure(0, weight=1)

		self.interf.grid_rowconfigure(1, weight=1)
		self.interf.grid_columnconfigure(1, weight=1)

		frame1 = Frame(self.interf)
		frame2 = Frame(self.interf)

		frame1.grid(row=0, column=0, sticky="nsew")
		frame2.grid(row=0, column=2, sticky="nsew")



		label1 = Label(frame1,text='Donor')
		load_file_d = partial(self.load_donor_file, frame1)
		label1.grid()


		b1 = Button(frame1, text="Browse Donor File", command=load_file_d, width=20)
		b1.grid()

		label2 = Label(frame2, text='Recipient')
		load_file_r = partial(self.load_recipient_file, frame2)
		label2.grid()


		b2 = Button(frame2, text="Browse Recipient File", command=load_file_r, width=20)
		b2.grid()



	def load_donor_file(self, frame):
		fname = askopenfilename(filetypes=(("Excel files", "*.xlsx"),
											("All files", "*.*") ))
		print(fname)
		if fname:

			name = os.path.basename(fname)
			ex_name = Label(frame,text=name)
			ex_name.grid()
			self.get_donor_sheets(fname, frame)

			return fname

	def get_donor_sheets(self, filename, frame):
		wb = openpyxl.load_workbook(filename)
		ws = wb.active
		sheets = wb.sheetnames
		print(sheets)

		variable = StringVar(self.interf)
		variable.set(sheets[0]) # default value


		self.donor_workbook = wb
		self.donor_sheet = wb[sheets[0]]
		w = OptionMenu(frame, variable, *sheets, command=self.d_sheet)
		w.grid()
		self.d_range_input(frame)

	def d_sheet(self, value):

		self.donor_sheet = self.donor_workbook[value]
		print(self.donor_sheet)

	def d_get_input(self, field):
		inp = field.get()
		print("get input")
		self.d_text_inp = inp
		print(self.d_text_inp)
		self.d_process_input()

	def d_range_input(self, frame):
		label = Label(frame, text='please enter range:')
		label.grid(sticky='w')
		entry = Entry(frame, bd = 5)
		entry.grid(sticky='e')

		get_user_input = partial(self.d_get_input, entry)

		input_button = Button(frame, text="submit", width=10, command=get_user_input)
		input_button.grid(sticky='e')
		print(self.donor_range)

	def d_process_input(self):
		print('in donor process input')

		vals = self.d_text_inp.split('-')
		ind = 0
		column = ""
		num = ""
		for char in vals[0]:
			if not char.isdigit():
				column += char
				ind += 1 
			else:
				num += char

		column = vals[0][0:ind]
		start = vals[0][ind:]
		end = vals[1][ind:]

	
		self.donor_range = {'column': column, 'start': start, 'end': end}
		self.range_eq()


	def load_recipient_file(self, frame):
		fname = askopenfilename(filetypes=(("Excel files", "*.xlsx"),
											("All files", "*.*") ))
		print(fname)
		if fname:
			name = os.path.basename(fname)
			ex_name = Label(frame,text=name)
			ex_name.grid()
			self.get_recipient_sheets(fname, frame)
			self.output_fname = fname

	def get_recipient_sheets(self, filename, frame):
		wb = openpyxl.load_workbook(filename)
		ws = wb.active
		sheets = wb.sheetnames
		print(sheets)

		variable = StringVar(self.interf)
		variable.set(sheets[0]) # default value


		self.recipient_workbook = wb
		self.recipient_sheet = wb[sheets[0]]
		g = OptionMenu(frame, variable, *sheets, command=self.r_sheet)
		g.grid()
		self.r_range_input(frame)

	def r_get_input(self, field):
		inp = field.get()
		print("get input")
		self.r_text_inp = inp
		print(self.r_text_inp)
		self.r_process_input()

	def r_process_input(self):
		print('in recipient process input')

		vals = self.r_text_inp.split('-')
		ind = 0
		column = ""
		num = ""
		for char in vals[0]:
			if not char.isdigit():
				column += char
				ind += 1 
			else:
				num += char

		column = vals[0][0:ind]
		start = vals[0][ind:]
		end = vals[1][ind:]

	
		self.recipient_range = {'column': column, 'start': start, 'end': end}
		self.range_eq()


	def r_range_input(self, frame):
		label = Label(frame, text='please enter range:')
		label.grid(sticky='w')
		entry = Entry(frame, bd = 5)
		entry.grid(sticky='e')

		get_user_input = partial(self.r_get_input, entry)

		input_button = Button(frame, text="submit", width=10, command=get_user_input)
		input_button.grid(sticky='e')
		print(self.donor_range)

	def r_sheet(self, value):
		self.recipient_sheet = self.recipient_workbook[value]
		print(self.recipient_sheet)


	def range_eq(self):
		if self.donor_range != None and self.recipient_range != None:
			d_diff = int(self.donor_range['end']) - int(self.donor_range['start'])
			r_diff = int(self.recipient_range['end']) - int(self.recipient_range['start'])

			distance = d_diff - r_diff

			if distance == 0:
				self.transfer()
				frame3 = Frame(self.interf)

				
				frame3.grid(row=0, column=1, sticky="nsew")
				label3 = Label(frame3,text='transfer begin')
				label3.grid()
			else: 
				frame3 = Frame(self.interf)

				
				frame3.grid(row=0, column=1, sticky="nsew")
				label3 = Label(frame3,text='range mismatch, please enter two ranges that are equal')
				label3.grid()


		else:
			print("not done yet")

	def transfer(self):
		donor = self.donor_sheet
		recipient = self.recipient_sheet

		d_s = self.donor_range['column'] + self.donor_range['start']
		d_e = self.donor_range['column'] + self.donor_range['end']

		r_s = self.recipient_range['column'] + self.recipient_range['start']
		r_e = self.recipient_range['column'] + self.recipient_range['end']

		values = []

		for i, rowOfCellObjects in enumerate(donor[d_s:d_e]):
			print(i)
			for n, cellObj in enumerate(rowOfCellObjects):
				print(cellObj.value)
				values.append(cellObj.value)

		for i, rowOfCellObjects in enumerate(recipient[r_s:r_e]):
			print(i)
			for n, cellObj in enumerate(rowOfCellObjects):
				cellObj.value = values[i]		
		self.recipient_workbook.save(self.output_fname)

	def launch(self):
		self.interf.mainloop()

if __name__ == "__main__":
	interface().launch()